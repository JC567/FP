# -*- coding: utf-8 -*-
"""高亮与满足判定 + 彩色 Excel 导出。

规则：
- 低位高亮：近10年PE百分位 < 30 的行 → 黄色
- 特殊高亮：K列(是否保留)==是 且 满足 L/M/N 之一（当前值 ≤ 提醒阈值，某列空则跳过）→ 绿色
  其中 L=买入提醒pe（用当前PE），M=买入提醒价格（用当前价/最新价），N=买入pb（用当前PB）
  特殊高亮优先级高于低位高亮。
当前 PE 从本地库 pe_hist 取最后值，当前 PB 从百度接口取市净率，结果缓存到 valuation_current 表。
"""
import os
import akshare as ak
import pandas as pd

import stock_db as db

DATA = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'data')


def _current_pb(conn, code):
    """取当前市净率（百度），失败返回 None。"""
    try:
        df = ak.stock_zh_valuation_baidu(symbol=code, indicator='市净率', period='近十年')
        v = df['value'].astype(float).dropna()
        return float(v.iloc[-1]) if not v.empty else None
    except Exception:
        return None


def get_pe_pb(conn, code):
    """返回 (当前PE, 当前PB)。PE 取本地 pe_hist 最后值；PB 取缓存或远程，并写回缓存。"""
    cur = db.get_valuation(conn, code)
    if cur is not None:
        return cur
    pe = None
    h = db.get_pe(conn, code)
    if h is not None and not h['value'].astype(float).dropna().empty:
        pe = float(h['value'].astype(float).dropna().iloc[-1])
    pb = _current_pb(conn, code)
    db.save_valuation(conn, code, pe, pb)
    return (pe, pb)


def compute_flags(df):
    """df 需含列：代码, 近10年PE百分位, 是否保留, 最新价, 当前PE, 当前PB,
    买入提醒pe, 买入提醒价格, 买入pb。
    返回 (pe_low: Series[bool], special: Series[bool])，索引与 df 一致。"""
    df = df.reset_index(drop=True)
    pe_low = df['近10年PE百分位'].notna() & (df['近10年PE百分位'] < 30)

    special = pd.Series(False, index=df.index)
    mask = df['是否保留'].astype(str).str.strip() == '是'
    for i in df.loc[mask].index:
        row = df.loc[i]
        hit = False
        L = row['买入提醒pe']; M = row['买入提醒价格']; N = row['买入pb']
        pe = row['当前PE']; pb = row['当前PB']
        if pd.notna(L) and pd.notna(pe) and pe <= L:
            hit = True
        if pd.notna(M) and pd.notna(row['最新价']) and row['最新价'] <= M:
            hit = True
        if pd.notna(N) and pd.notna(pb) and pb <= N:
            hit = True
        special.at[i] = hit
    return pe_low, special


def export_excel(df, pe_low, special, out_path):
    """导出带填充色的 Excel：低位=黄，满足=绿（优先）。"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = '分红率排名'
    cols = list(df.columns)
    ws.append(cols)

    yellow = PatternFill('solid', fgColor='FFF2CC')
    green = PatternFill('solid', fgColor='C6EFCE')
    head_font = Font(bold=True)
    for c in range(1, len(cols) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = head_font
        cell.fill = PatternFill('solid', fgColor='DDEBF7')
        cell.alignment = Alignment(horizontal='center')

    for i, (_, row) in enumerate(df.iterrows(), start=2):
        fill = None
        if special.iloc[i - 2]:
            fill = green
        elif pe_low.iloc[i - 2]:
            fill = yellow
        for c, col in enumerate(cols, 1):
            cell = ws.cell(row=i, column=c, value=row[col])
            if fill:
                cell.fill = fill

    for c, col in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(c)].width = min(
            max(len(str(col)) * 2 + 6, 12), 22)
    ws.freeze_panes = 'A2'
    wb.save(out_path)


def process(csv_path, xlsx_path, label=''):
    """读 CSV → 计算标记 → 导出彩色 Excel。返回 (低位数, 满足数)。"""
    df = pd.read_csv(csv_path, encoding='utf-8-sig')
    df['代码'] = df['代码'].astype(str).str.replace(r'\D', '', regex=True).str.zfill(6)
    pe_low, special = compute_flags(df)
    export_excel(df, pe_low, special, xlsx_path)
    print(f'{label} 高亮完成: PE低位(PE<30)={int(pe_low.sum())} 只, 满足提醒(K=是且L/M/N任一达标)={int(special.sum())} 只 -> {xlsx_path}')
    return int(pe_low.sum()), int(special.sum())


if __name__ == '__main__':
    process(os.path.join(DATA, '分红率排名（筛选后）.csv'),
            os.path.join(DATA, '分红率排名（筛选后）.xlsx'), '筛选后')
